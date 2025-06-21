"""
File Operations Functions
"""

import os
import json
import csv
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional
import openpyxl
from .base import BaseFunction


class ReadFileFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "read_file"
    
    @property
    def description(self) -> str:
        return "Read content from a text file"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, file_path: str, encoding: str = "utf-8") -> Dict[str, Any]:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                content = file.read()
            
            return {
                "success": True,
                "content": content,
                "size": len(content),
                "lines": len(content.splitlines())
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


class WriteFileFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "write_file"
    
    @property
    def description(self) -> str:
        return "Write content to a text file"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, file_path: str, content: str, mode: str = "w", encoding: str = "utf-8") -> Dict[str, Any]:
        try:
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            with open(file_path, mode, encoding=encoding) as file:
                file.write(content)
            
            return {
                "success": True,
                "message": f"Content written to {file_path}",
                "bytes_written": len(content.encode(encoding))
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


class CopyFileFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "copy_file"
    
    @property
    def description(self) -> str:
        return "Copy a file from source to destination"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, source_path: str, destination_path: str) -> Dict[str, Any]:
        try:
            # Create destination directory if it doesn't exist
            os.makedirs(os.path.dirname(destination_path), exist_ok=True)
            
            shutil.copy2(source_path, destination_path)
            
            return {
                "success": True,
                "message": f"File copied from {source_path} to {destination_path}"
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


class DeleteFileFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "delete_file"
    
    @property
    def description(self) -> str:
        return "Delete a file"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, file_path: str) -> Dict[str, Any]:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                return {"success": True, "message": f"File {file_path} deleted"}
            else:
                return {"success": False, "error": f"File {file_path} not found"}
        except Exception as e:
            return {"success": False, "error": str(e)}


class ListDirectoryFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "list_directory"
    
    @property
    def description(self) -> str:
        return "List files and directories in a path"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, directory_path: str, include_hidden: bool = False) -> Dict[str, Any]:
        try:
            if not os.path.exists(directory_path):
                return {"success": False, "error": f"Directory {directory_path} not found"}
            
            items = []
            for item in os.listdir(directory_path):
                if not include_hidden and item.startswith('.'):
                    continue
                
                item_path = os.path.join(directory_path, item)
                items.append({
                    "name": item,
                    "type": "directory" if os.path.isdir(item_path) else "file",
                    "size": os.path.getsize(item_path) if os.path.isfile(item_path) else None,
                    "modified": os.path.getmtime(item_path)
                })
            
            return {"success": True, "items": items, "count": len(items)}
        except Exception as e:
            return {"success": False, "error": str(e)}


class CreateDirectoryFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "create_directory"
    
    @property
    def description(self) -> str:
        return "Create a new directory"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, directory_path: str, parents: bool = True) -> Dict[str, Any]:
        try:
            if parents:
                os.makedirs(directory_path, exist_ok=True)
            else:
                os.mkdir(directory_path)
            
            return {"success": True, "message": f"Directory {directory_path} created"}
        except Exception as e:
            return {"success": False, "error": str(e)}


class ReadJSONFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "read_json"
    
    @property
    def description(self) -> str:
        return "Read and parse JSON file"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, file_path: str) -> Dict[str, Any]:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            return {"success": True, "data": data}
        except Exception as e:
            return {"success": False, "error": str(e)}


class WriteJSONFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "write_json"
    
    @property
    def description(self) -> str:
        return "Write data to JSON file"
    
    @property
    def category(self) -> str:
        return "file_operations"
    
    async def execute(self, file_path: str, data: Any, indent: int = 2) -> Dict[str, Any]:
        try:
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump(data, file, indent=indent, ensure_ascii=False)
            
            return {"success": True, "message": f"Data written to {file_path}"}
        except Exception as e:
            return {"success": False, "error": str(e)}


class ReadExcelFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "read_excel"

    @property
    def description(self) -> str:
        return "Read data from Excel file"

    @property
    def category(self) -> str:
        return "file_operations"

    async def execute(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        try:
            workbook = openpyxl.load_workbook(file_path)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            data = []
            headers = []

            for row_num, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if row_num == 1:
                    headers = [cell or f"Column_{i+1}" for i, cell in enumerate(row)]
                else:
                    row_data = dict(zip(headers, row))
                    data.append(row_data)

            return {
                "success": True,
                "data": data,
                "sheet_name": worksheet.title,
                "rows": len(data),
                "columns": len(headers)
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


class GetFileInfoFunction(BaseFunction):
    @property
    def name(self) -> str:
        return "get_file_info"

    @property
    def description(self) -> str:
        return "Get detailed information about a file"

    @property
    def category(self) -> str:
        return "file_operations"

    async def execute(self, file_path: str) -> Dict[str, Any]:
        try:
            if not os.path.exists(file_path):
                return {"success": False, "error": f"File {file_path} not found"}

            stat = os.stat(file_path)

            return {
                "success": True,
                "info": {
                    "name": os.path.basename(file_path),
                    "path": os.path.abspath(file_path),
                    "size": stat.st_size,
                    "created": stat.st_ctime,
                    "modified": stat.st_mtime,
                    "accessed": stat.st_atime,
                    "is_file": os.path.isfile(file_path),
                    "is_directory": os.path.isdir(file_path),
                    "extension": os.path.splitext(file_path)[1]
                }
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
